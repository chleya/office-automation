"""
Excel Spreadsheet Processor

This module provides functionality for creating, editing, and manipulating
Excel spreadsheets (.xlsx format) with WPS compatibility.
"""

import os
from typing import Optional, List, Dict, Any, Union, Tuple
from pathlib import Path
import warnings

# Try to import error_handler with different methods
try:
    # First try absolute import
    from utils.error_handler import (
        DocumentCreationError,
        DocumentReadError,
        DocumentSaveError,
        ValidationError,
        create_error_context,
        validate_file_path,
        validate_format,
    )
except ImportError:
    try:
        # Try relative import
        from ..utils.error_handler import (
            DocumentCreationError,
            DocumentReadError,
            DocumentSaveError,
            ValidationError,
            create_error_context,
            validate_file_path,
            validate_format,
        )
    except ImportError:
        # Create dummy error classes
        class DocumentCreationError(Exception):
            pass
        
        class DocumentReadError(Exception):
            pass
        
        class DocumentSaveError(Exception):
            pass
        
        class ValidationError(Exception):
            pass
        
        def create_error_context(*args, **kwargs):
            return None
        
        def validate_file_path(*args, **kwargs):
            return True
        
        def validate_format(*args, **kwargs):
            return True


class ExcelProcessor:
    """Processor for Excel spreadsheet operations."""
    
    # Supported formats
    SUPPORTED_FORMATS = ["xlsx", "xls", "csv", "pdf", "html"]
    
    def __init__(self, template_dir: Optional[str] = None):
        """
        Initialize Excel Processor.
        
        Args:
            template_dir: Directory containing Excel templates
        """
        self.template_dir = template_dir
        self._workbook = None
        self._active_worksheet = None
        
        # Try to import openpyxl
        try:
            import openpyxl
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
            from openpyxl.drawing.image import Image
            from openpyxl.utils import get_column_letter
            
            self.Workbook = Workbook
            self.load_workbook = load_workbook
            self.Font = Font
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
            self.Alignment = Alignment
            self.BarChart = BarChart
            self.LineChart = LineChart
            self.PieChart = PieChart
            self.ScatterChart = ScatterChart
            self.Image = Image
            self.get_column_letter = get_column_letter
            
            self._openpyxl_available = True
        except ImportError:
            self._openpyxl_available = False
            raise ImportError(
                "openpyxl is not installed. Please install it with: "
                "pip install openpyxl"
            )
    
    def create_workbook(
        self,
        template: Optional[Union[str, Path]] = None
    ) -> 'Workbook':
        """
        Create a new Excel workbook.
        
        Args:
            template: Path to template file or template name
            
        Returns:
            Workbook object
            
        Raises:
            DocumentCreationError: If workbook creation fails
        """
        context = create_error_context("create_workbook", template=template)
        
        try:
            if template:
                # Load from template
                template_path = self._resolve_template_path(template)
                if not os.path.exists(template_path):
                    raise FileNotFoundError(f"Template not found: {template_path}")
                
                self._workbook = self.load_workbook(str(template_path))
            else:
                # Create empty workbook
                self._workbook = self.Workbook()
                # Remove default sheet if it's empty
                if len(self._workbook.sheetnames) == 1:
                    default_sheet = self._workbook.active
                    if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                        self._workbook.remove(default_sheet)
            
            # Set first sheet as active
            if self._workbook.sheetnames:
                self._active_worksheet = self._workbook[self._workbook.sheetnames[0]]
            else:
                self._active_worksheet = None
            
            return self._workbook
            
        except Exception as e:
            raise DocumentCreationError(
                f"Failed to create workbook: {e}",
                context
            )
    
    def load_workbook_file(self, file_path: Union[str, Path]) -> 'Workbook':
        """
        Load an existing Excel workbook.
        
        Args:
            file_path: Path to the workbook file
            
        Returns:
            Workbook object
            
        Raises:
            DocumentReadError: If workbook loading fails
        """
        context = create_error_context("load_workbook", file_path=file_path)
        validate_file_path(str(file_path), "load_workbook")
        
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"Workbook not found: {file_path}")
            
            self._workbook = self.load_workbook(str(file_path))
            
            # Set first sheet as active
            if self._workbook.sheetnames:
                self._active_worksheet = self._workbook[self._workbook.sheetnames[0]]
            else:
                self._active_worksheet = None
            
            return self._workbook
            
        except Exception as e:
            raise DocumentReadError(
                f"Failed to load workbook: {e}",
                context
            )
    
    def add_worksheet(
        self,
        name: str,
        data: Optional[List[List[Any]]] = None,
        headers: Optional[List[str]] = None
    ) -> 'Worksheet':
        """
        Add a new worksheet to the workbook.
        
        Args:
            name: Worksheet name
            data: Initial data for the worksheet
            headers: Column headers
            
        Returns:
            Worksheet object
            
        Raises:
            DocumentCreationError: If worksheet creation fails
        """
        if not self._workbook:
            raise DocumentCreationError(
                "Workbook not initialized. Call create_workbook() or load_workbook_file() first."
            )
        
        try:
            # Create new worksheet
            worksheet = self._workbook.create_sheet(title=name)
            
            # Set as active worksheet
            self._active_worksheet = worksheet
            
            # Add headers if provided
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=1, column=col_idx, value=header)
                    cell.font = self.Font(bold=True)
            
            # Add data if provided
            if data:
                start_row = 2 if headers else 1
                for row_idx, row_data in enumerate(data, start=start_row):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            return worksheet
            
        except Exception as e:
            raise DocumentCreationError(f"Failed to add worksheet: {e}")
    
    def set_active_worksheet(self, name_or_index: Union[str, int]) -> 'Worksheet':
        """
        Set the active worksheet.
        
        Args:
            name_or_index: Worksheet name or index (0-based)
            
        Returns:
            Worksheet object
            
        Raises:
            ValueError: If worksheet not found
        """
        if not self._workbook:
            raise DocumentCreationError(
                "Workbook not initialized. Call create_workbook() or load_workbook_file() first."
            )
        
        try:
            if isinstance(name_or_index, str):
                # Find by name
                if name_or_index in self._workbook.sheetnames:
                    self._active_worksheet = self._workbook[name_or_index]
                else:
                    raise ValueError(f"Worksheet '{name_or_index}' not found")
            else:
                # Find by index
                if 0 <= name_or_index < len(self._workbook.sheetnames):
                    sheet_name = self._workbook.sheetnames[name_or_index]
                    self._active_worksheet = self._workbook[sheet_name]
                else:
                    raise ValueError(f"Worksheet index {name_or_index} out of range")
            
            return self._active_worksheet
            
        except Exception as e:
            raise DocumentCreationError(f"Failed to set active worksheet: {e}")
    
    def set_cell_value(
        self,
        cell_ref: str,
        value: Any,
        formula: Optional[str] = None
    ) -> None:
        """
        Set value or formula for a cell.
        
        Args:
            cell_ref: Cell reference (e.g., "A1", "B2")
            value: Cell value
            formula: Cell formula (starts with "=")
            
        Raises:
            DocumentCreationError: If no active worksheet
        """
        if not self._active_worksheet:
            raise DocumentCreationError(
                "No active worksheet. Call set_active_worksheet() first."
            )
        
        try:
            cell = self._active_worksheet[cell_ref]
            
            if formula:
                cell.value = formula
            else:
                cell.value = value
                
        except Exception as e:
            raise DocumentCreationError(f"Failed to set cell value: {e}")
    
    def get_cell_value(self, cell_ref: str) -> Any:
        """
        Get value from a cell.
        
        Args:
            cell_ref: Cell reference (e.g., "A1", "B2")
            
        Returns:
            Cell value
            
        Raises:
            DocumentCreationError: If no active worksheet
        """
        if not self._active_worksheet:
            raise DocumentCreationError(
                "No active worksheet. Call set_active_worksheet() first."
            )
        
        try:
            return self._active_worksheet[cell_ref].value
        except Exception as e:
            raise DocumentCreationError(f"Failed to get cell value: {e}")
    
    def add_chart(
        self,
        data_range: str,
        chart_type: str = "bar",
        title: Optional[str] = None,
        position: str = "A1"
    ) -> None:
        """
        Add a chart to the worksheet.
        
        Args:
            data_range: Data range (e.g., "A1:B10")
            chart_type: Type of chart ("bar", "line", "pie", "scatter")
            title: Chart title
            position: Position to place the chart (e.g., "D1")
            
        Raises:
            DocumentCreationError: If no active worksheet
        """
        if not self._active_worksheet:
            raise DocumentCreationError(
                "No active worksheet. Call set_active_worksheet() first."
            )
        
        try:
            # Create chart based on type
            if chart_type.lower() == "bar":
                chart = self.BarChart()
            elif chart_type.lower() == "line":
                chart = self.LineChart()
            elif chart_type.lower() == "pie":
                chart = self.PieChart()
            elif chart_type.lower() == "scatter":
                chart = self.ScatterChart()
            else:
                raise ValueError(f"Unsupported chart type: {chart_type}")
            
            # Set chart title
            if title:
                chart.title = title
            
            # Add data
            # Note: This is simplified - in production you'd parse the data_range
            # and create proper data references
            
            # Add chart to worksheet
            self._active_worksheet.add_chart(chart, position)
            
        except Exception as e:
            raise DocumentCreationError(f"Failed to add chart: {e}")
    
    def apply_formatting(
        self,
        cell_range: str,
        format_config: Dict[str, Any]
    ) -> None:
        """
        Apply formatting to a cell range.
        
        Args:
            cell_range: Cell range (e.g., "A1:B10")
            format_config: Formatting configuration
            
        Raises:
            DocumentCreationError: If no active worksheet
        """
        if not self._active_worksheet:
            raise DocumentCreationError(
                "No active worksheet. Call set_active_worksheet() first."
            )
        
        try:
            # Parse cell range (simplified)
            # In production, you'd use openpyxl's range parsing
            
            # Apply formatting based on config
            # This is a placeholder - actual implementation would
            # apply font, fill, border, alignment, etc.
            
            warnings.warn(
                "Formatting implementation is simplified. "
                "Full formatting support requires more complex implementation."
            )
            
        except Exception as e:
            raise DocumentCreationError(f"Failed to apply formatting: {e}")
    
    def save(
        self,
        file_path: Union[str, Path],
        format: str = "xlsx"
    ) -> str:
        """
        Save the workbook to a file.
        
        Args:
            file_path: Path where to save the workbook
            format: Output format (xlsx, csv, etc.)
            
        Returns:
            Path to saved file
            
        Raises:
            DocumentSaveError: If saving fails
        """
        context = create_error_context(
            "save_workbook",
            file_path=file_path,
            format=format
        )
        
        validate_file_path(str(file_path), "save")
        validate_format(format, self.SUPPORTED_FORMATS, "save")
        
        if not self._workbook:
            raise DocumentSaveError(
                "No workbook to save. Create or load a workbook first.",
                context
            )
        
        try:
            file_path = Path(file_path)
            
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Save in requested format
            if format == "xlsx":
                self._workbook.save(str(file_path))
            elif format == "csv":
                self._save_as_csv(file_path)
            elif format == "pdf":
                self._save_as_pdf(file_path)
            else:
                # Default to xlsx
                self._workbook.save(str(file_path))
            
            return str(file_path)
            
        except Exception as e:
            raise DocumentSaveError(
                f"Failed to save workbook: {e}",
                context
            )
    
    def _save_as_csv(self, file_path: Path) -> None:
        """Save active worksheet as CSV."""
        if not self._active_worksheet:
            raise DocumentSaveError("No active worksheet to save as CSV")
        
        import csv
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write all rows
            for row in self._active_worksheet.iter_rows(values_only=True):
                writer.writerow(row)
    
    def _save_as_pdf(self, file_path: Path) -> None:
        """Save workbook as PDF."""
        # This is a placeholder - in production, you'd use a proper PDF converter
        raise NotImplementedError(
            "PDF export requires additional setup. "
            "Install openpyxl with PDF support or use a separate PDF library."
        )
    
    def _resolve_template_path(self, template: Union[str, Path]) -> Path:
        """Resolve template path."""
        template = str(template)
        
        # If it's an absolute path or relative path, use as-is
        if os.path.isabs(template) or template.startswith('.'):
            return Path(template)
        
        # Check in template directory
        if self.template_dir:
            template_path = Path(self.template_dir) / template
            if template_path.exists():
                return template_path
        
        # Check common extensions
        for ext in ['.xlsx', '.xls', '.xltx']:
            test_path = Path(template + ext)
            if test_path.exists():
                return test_path
        
        # Return as-is (will fail if not found)
        return Path(template)
    
    def get_workbook(self) -> Optional['Workbook']:
        """Get the current workbook object."""
        return self._workbook
    
    def get_active_worksheet(self) -> Optional['Worksheet']:
        """Get the active worksheet object."""
        return self._active_worksheet
    
    def clear(self) -> None:
        """Clear the current workbook."""
        self._workbook = None
        self._active_worksheet = None
    
    def is_initialized(self) -> bool:
        """Check if workbook is initialized."""
        return self._workbook is not None
    
    def get_worksheet_names(self) -> List[str]:
        """Get list of worksheet names."""
        if not self._workbook:
            return []
        return self._workbook.sheetnames
    
    def get_worksheet(self, name_or_index: Union[str, int]) -> 'Worksheet':
        """Get worksheet by name or index."""
        if not self._workbook:
            raise DocumentCreationError("Workbook not initialized")
        
        if isinstance(name_or_index, str):
            if name_or_index in self._workbook.sheetnames:
                return self._workbook[name_or_index]
            else:
                raise ValueError(f"Worksheet '{name_or_index}' not found")
        else:
            if 0 <= name_or_index < len(self._workbook.sheetnames):
                sheet_name = self._workbook.sheetnames[name_or_index]
                return self._workbook[sheet_name]
            else:
                raise ValueError(f"Worksheet index {name_or_index} out of range")